"""Spreadsheet decomposer with full-length string chunking."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .base import BaseDecomposer
from semantic_types import chunk_str, is_chunk_key, rejoin_chunks, CHUNK_SIZE

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


class XlsxDecomposer(BaseDecomposer):
    def _is_csv(self) -> bool:
        return Path(self.file_path).suffix.lower() == ".csv"

    # ------------------------------------------------------------------
    # to_entries
    # ------------------------------------------------------------------

    def to_entries(self) -> list[tuple[str, Any]]:
        if self._is_csv():
            return self._csv_to_entries()
        return self._xlsx_to_entries()

    def _encode_cell(self, key: str, value: Any) -> list[tuple[str, Any]]:
        """Return one or more entries for a cell value.

        - Numeric / bool values fit in one pixel.
        - Strings longer than CHUNK_SIZE are split into chunk entries.
        - Short strings (<= CHUNK_SIZE chars) are stored as a single entry.
        """
        if isinstance(value, bool):
            return [(key, value)]
        if isinstance(value, (int, float)):
            return [(key, value)]
        text = str(value)
        if len(text) <= CHUNK_SIZE:
            return [(key, text)]
        return chunk_str(key, text)

    def _csv_to_entries(self) -> list[tuple[str, Any]]:
        entries: list[tuple[str, Any]] = [("meta.original_extension", "csv")]
        with open(self.file_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader, start=1):
                for c, value in enumerate(row, start=1):
                    if value != "":
                        for entry in self._encode_cell(f"R{r}C{c}", value):
                            entries.append(entry)
        return entries

    def _xlsx_to_entries(self) -> list[tuple[str, Any]]:
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is required for XLSX support")
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        entries: list[tuple[str, Any]] = [("meta.original_extension", "xlsx")]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_key = f"{sheet_name}!R{cell.row}C{cell.column}"
                        for entry in self._encode_cell(cell_key, cell.value):
                            entries.append(entry)
        return entries

    # ------------------------------------------------------------------
    # from_entries
    # ------------------------------------------------------------------

    def from_entries(self, entries: list[tuple[str, Any]], output_path: str) -> None:
        ext = Path(output_path).suffix.lower()
        if ext == ".csv":
            self._entries_to_csv(entries, output_path)
        else:
            self._entries_to_xlsx(entries, output_path)

    def _resolve_entries(self, entries: list[tuple[str, Any]]) -> dict[str, Any]:
        """Collapse chunk entries back into full string values.

        Pass 1: build raw lookup.
        Pass 2: for every key that has a __chunk_count, rejoin and store
                under the original key; skip all __chunk.* entries.
        """
        raw: dict[str, Any] = dict(entries)

        resolved: dict[str, Any] = {}
        # Collect base keys that were chunked
        chunked_bases: set[str] = set()
        for key in raw:
            if key.endswith(".__chunk_count"):
                base = key[: -len(".__chunk_count")]
                chunked_bases.add(base)

        for key, value in raw.items():
            if is_chunk_key(key):
                continue  # handled separately
            if key in chunked_bases:
                # This would be the chunk_count entry — skip, handled below
                continue
            resolved[key] = value

        for base in chunked_bases:
            resolved[base] = rejoin_chunks(base, raw)

        return resolved

    def _entries_to_csv(self, entries: list[tuple[str, Any]], output_path: str) -> None:
        lookup = self._resolve_entries(entries)
        grid: dict[tuple[int, int], str] = {}
        for key, value in lookup.items():
            if not (key.startswith("R") and "C" in key):
                continue
            try:
                r_str, c_str = key[1:].split("C", 1)
                grid[(int(r_str), int(c_str))] = str(value)
            except ValueError:
                continue

        max_row = max((r for r, _ in grid), default=0)
        max_col = max((c for _, c in grid), default=0)
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for r in range(1, max_row + 1):
                writer.writerow([grid.get((r, c), "") for c in range(1, max_col + 1)])

    def _entries_to_xlsx(self, entries: list[tuple[str, Any]], output_path: str) -> None:
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is required for XLSX support")

        lookup = self._resolve_entries(entries)

        wb = openpyxl.Workbook()
        # Remove the default empty sheet that openpyxl always creates
        default_sheet = wb.active
        wb.remove(default_sheet)

        ws_map: dict[str, Any] = {}

        for key, value in lookup.items():
            if "!R" not in key or "C" not in key:
                continue
            try:
                sheet_name, rc = key.split("!", 1)
                r_str, c_str = rc[1:].split("C", 1)
                row, col = int(r_str), int(c_str)
            except ValueError:
                continue

            if sheet_name not in ws_map:
                ws_map[sheet_name] = wb.create_sheet(title=sheet_name)
            ws_map[sheet_name].cell(row=row, column=col, value=value)

        wb.save(output_path)
